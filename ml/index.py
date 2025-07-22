from email.mime import image
import cv2
import numpy as np
from twilio.rest import Client
from ultralytics import YOLO
import boto3
from botocore.exceptions import NoCredentialsError
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# c:\Users\SINGAPORE\Desktop\Secure-Sight-main\your_script.py

# c:\Users\SINGAPORE\Desktop\Secure-Sight-main\anomaly_viewer.py

import os
from matplotlib import pyplot as plt
import cv2

anomaly_dir = 'path_to_anomaly_images'  # Update with your actual path
# No import needed for video link here. Just set the path as a string.

# List all image files in the anomaly directory
image_files = [f for f in os.listdir(anomaly_dir) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]

for img_file in image_files:
    img_path = os.path.join(anomaly_dir, img_file)
    img = cv2.imread(img_path)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    plt.imshow(img)
    plt.title(img_file)
    plt.axis('off')
    plt.show()

import cv2

# Assuming 'image' is your image array (e.g., from cv2.imread or a camera)
cv2.imwrite('output.png', image)  # Saves the image as PNG

# ...existing code...

# ...existing code...

cv2.imshow('Image', image)
cv2.waitKey(0)
cv2.destroyAllWindows()

# ...existing code...

# Load a model
camera_label = "Alert: Security Concern Detected Near Main Entrance!"
account_sid = 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa'
auth_token = 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa'
twilio_phone_number = 'whatsapp:+199999999999'
recipient_phone_numbers = ['whatsapp:+999999999999']

client = Client(account_sid, auth_token)

# Upload to S3 function
def upload_to_s3(file_name, bucket='ak-hackathon-bucket', object_name=None):
    s3_client = boto3.client('s3')
    try:
        s3_client.upload_file(
            Filename=file_name,
            Bucket=bucket,
            Key=object_name or file_name,
            ExtraArgs={'ContentType': 'image/jpeg'}
        )
        url = f'https://{bucket}.s3.amazonaws.com/{object_name or file_name}'
        return url
    except FileNotFoundError:
        print(f"The file {file_name} was not found.")
        return None
    except NoCredentialsError:
        print("Credentials not available.")
        return None
    except Exception as e:
        print(f"Error uploading to S3: {e}")
        return None

# Send alert with image URL
def send_whatsapp_alert(alert_message, to, media_url=None):
    try:
        message = client.messages.create(
            to=to,
            from_=twilio_phone_number,
            body=alert_message,
            media_url=[media_url] if media_url else None
        )
        print("Message SID:", message.sid)
    except Exception as e:
        print(f"Error sending WhatsApp message: {e}")

# Save alert information to Excel
def save_alert_to_excel(timestamp, alert_type, file_path, excel_path='alerts.xlsx'):
    img_path = os.path.abspath(file_path)
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(['Timestamp', 'Alert Type', 'Image Path'])
        wb.save(excel_path)
    wb = load_workbook(excel_path)
    ws = wb.active
    ws.append([timestamp, alert_type, img_path])
    wb.save(excel_path)

# Set up YOLO model
model_path = "/Users/apple/Desktop/MakeUC2024/models/best_final_40.pt"
model = YOLO(model_path)

# Directory to store anomaly images
anomaly_img_dir = os.path.join(os.path.dirname(__file__), "anomaly_images")
if not os.path.exists(anomaly_img_dir):
    os.makedirs(anomaly_img_dir, exist_ok=True)

# Directory to store alert images
alert_img_dir = os.path.join(os.path.dirname(model_path), "best_final_40.pt")
if not os.path.exists(alert_img_dir):
    os.makedirs(alert_img_dir, exist_ok=True)

# Video capture setup
# video_path = "/Users/apple/Desktop/MakeUC2024/demo_videos/Filipino Knife Defense is Deadly!!.mp4"
# video_path="/Users/apple/Desktop/MakeUC2024/demo_videos/demo2.mp4"
video_path="/Users/apple/Desktop/MakeUC2024/demo_videos/Extraction ï½ One-Shot Gun Fight Scene.mp4"
cap = cv2.VideoCapture(video_path)


# cap=cv2.VideoCapture(0)
# Alert tracking and cooldown settings
alert_cooldown = 45  
last_alert_time = {"gun": 0, "knife": 0}

# Main loop for real-time detection
while cap.isOpened():
    ret, frame = cap.read()
    if not ret:
        print("Error: Failed to grab a frame.")
        break

    # Optionally, reduce frame size to improve FPS
    frame = cv2.resize(frame, (640, 360))

    # Run YOLO detection on the frame
    results = model(frame)

    # Get the current timestamp for cooldown checking
    current_time = time.time()

    # Process the results
    for r in results:
        if r.boxes:
            # Check detected classes and prepare alerts accordingly
            for box in r.boxes:
                class_id = int(box.cls)
                x1, y1, x2, y2 = map(int, box.xyxy[0])  # Get box coordinates

                # Draw bounding box
                color = (0, 255, 0) if class_id == 2 else (0, 0, 255)  # Green for person, red for weapons
                cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)

                # Add label
                label = "Person" if class_id == 2 else "Gun" if class_id == 0 else "Knife"
                cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

                # Determine if an alert should be sent
                if class_id == 0 and (current_time - last_alert_time["gun"] > alert_cooldown):
                    alert_message = f"{camera_label} A potential threat has been detected involving a firearm near university area. Stay alert and avoid the vicinity if possible. Security is being notified."
                    last_alert_time["gun"] = current_time
                    alert_type = "Gun"
                elif class_id == 1 and (current_time - last_alert_time["knife"] > alert_cooldown):
                    alert_message = f"{camera_label} A potential threat has been detected involving a near university area. Stay alert and avoid the vicinity if possible. Security is being notified."                    
                    last_alert_time["knife"] = current_time
                    alert_type = "Knife"
                else:
                    continue  # Skip if within cooldown period

                # Save the anomaly frame as PNG in anomaly_images directory
                file_name = f"anomaly_{datetime.now().strftime('%Y_%m_%d_%H_%M_%S')}.png"
                file_path = os.path.join(anomaly_img_dir, file_name)
                cv2.imwrite(file_path, frame)

                # Save alert info to Excel
                save_alert_to_excel(datetime.now().strftime('%Y-%m-%d %H:%M:%S'), alert_type, file_path)

                # Upload to S3 and get URL
                media_url = upload_to_s3(file_path)

                # Send WhatsApp alert if media_url is available
                if media_url:
                    for recipient in recipient_phone_numbers:
                        send_whatsapp_alert(alert_message, recipient, media_url=media_url)

    # Display the frame for visual feedback
    cv2.imshow('Object Detection', frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release resources
cap.release() 
cv2.destroyAllWindows()

# All detected anomaly images are saved in anomaly_images/ and logged in alerts.xlsx
